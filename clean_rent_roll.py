import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE = "sample_rent_roll.xlsx"
SHEET_NAME = 0
OUTPUT_FILE = "Rent_Roll_Organized.xlsx"

KNOWN_CHARGE_ORDER = [
    "Amount",
    "outpark",
    "vwaste",
    "packlock",
    "insclear",
    "amenityf",
    "rcable",
    "petrent",
    "rp_depC",
    "rp_depA",
    "mtmfee",
    "rentcon",
    "rempdisc",
    "pestctrl",
]

BASE_RENT_CODES = {"rrent", "rent", "renl", "rentl", "baserent", "base"}

CODE_COLUMN_NAME_MAP = {
    "rp_depc": "rp_depC",
    "rp_depa": "rp_depA",
}


def norm_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return "" if text.lower() == "nan" else text


def find_header_row(raw_df, max_scan=300):
    for i in range(min(len(raw_df), max_scan)):
        row = raw_df.iloc[i].astype(str).str.replace("\n", " ").str.strip().str.lower()
        vals = set(row.values)
        if "unit" in vals and "resident" in vals and "amount" in vals and ("charge" in vals or "rent code" in vals):
            return i
    return None


def has_two_row_header(raw_df, header_row):
    if header_row + 1 >= len(raw_df):
        return False
    next_row = raw_df.iloc[header_row + 1].astype(str).str.strip().str.lower().tolist()
    markers = {"sq ft", "rent", "code", "deposit", "expiration"}
    return any(v in markers for v in next_row)


def combine_header_name(top, sub):
    top = norm_text(top)
    sub = norm_text(sub)
    top_l = top.lower()
    sub_l = sub.lower()

    if top_l == "unit" and sub_l == "sq ft":
        return "Sq Ft"
    if top_l == "market" and sub_l == "rent":
        return "Market"
    if top_l == "charge" and sub_l == "code":
        return "Charge Code"
    if top_l == "resident" and sub_l == "deposit":
        return "Resident Deposit"
    if top_l == "other" and sub_l == "deposit":
        return "Other Deposit"
    if top_l == "lease" and sub_l == "expiration":
        return "Lease Expiration"
    return top if top else sub


def build_column_names(raw_df, header_row, two_row_header):
    top = raw_df.iloc[header_row]
    sub = raw_df.iloc[header_row + 1] if two_row_header else pd.Series([None] * len(top))

    names = [combine_header_name(a, b) for a, b in zip(top, sub)]
    while names and names[-1] == "":
        names.pop()
    return names


def canonical_unit(value):
    if pd.isna(value):
        return pd.NA
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return pd.NA
    n = pd.to_numeric(pd.Series([s]), errors="coerce").iloc[0]
    if pd.notna(n):
        if abs(n - round(n)) <= 1e-9:
            return str(int(round(n)))
        return f"{n:.6f}".rstrip("0").rstrip(".")
    return s


def first_non_null(series):
    s = series.dropna()
    return s.iloc[0] if len(s) else pd.NA


def normalize_charge_code(code):
    s = norm_text(code).lower()
    s = re.sub(r"\s+", "", s)
    if not s or s == "total":
        return None
    if s in BASE_RENT_CODES:
        return "Amount"
    return CODE_COLUMN_NAME_MAP.get(s, s)


def format_excel(path, date_columns):
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header is not None:
            header_to_col[str(header).strip()] = col_idx

    for date_col in date_columns:
        col_idx = header_to_col.get(date_col)
        if not col_idx:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.number_format = "mm/dd/yyyy"

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    wb.save(path)


raw = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=None)
header_row = find_header_row(raw)
if header_row is None:
    raise ValueError("Could not detect header row. The rent roll layout may have changed.")

two_row_header = has_two_row_header(raw, header_row)
column_names = build_column_names(raw, header_row, two_row_header)
data_start = header_row + (2 if two_row_header else 1)

df = pd.read_excel(
    INPUT_FILE,
    sheet_name=SHEET_NAME,
    header=None,
    skiprows=data_start,
    names=column_names,
    usecols=list(range(len(column_names))),
)

rename_map = {
    "Unit.1": "Sq Ft",
    "Market Rent": "Market",
    "Market": "Market",
    "Charge": "Charge Code",
    "Rent Code": "Charge Code",
    "Resident.1": "Resident Deposit",
    "Other": "Other Deposit",
    "Lease": "Lease Expiration",
}
df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

required_columns = ["Unit", "Charge Code", "Amount"]
missing_required = [c for c in required_columns if c not in df.columns]
if missing_required:
    raise ValueError(f"Missing required columns: {missing_required}")

meta_columns = [
    "Unit",
    "Unit Type",
    "Sq Ft",
    "Resident",
    "Name",
    "Market",
    "Resident Deposit",
    "Other Deposit",
    "Move In",
    "Lease Expiration",
    "Move Out",
    "Balance",
]

for col in meta_columns:
    if col not in df.columns:
        df[col] = pd.NA

df["_unit_start"] = df["Unit"].notna()
unit_meta = df[df["_unit_start"]].copy()
unit_meta["Unit"] = unit_meta["Unit"].apply(canonical_unit)
unit_meta = unit_meta[unit_meta["Unit"].notna()].copy()
unit_meta = unit_meta.groupby("Unit", as_index=False).agg(first_non_null)
unit_meta = unit_meta[meta_columns]

df["Unit"] = df["Unit"].ffill().apply(canonical_unit)
df = df[df["Unit"].notna()].copy()

charges = df[["Unit", "Charge Code", "Amount"]].copy()
charges["Charge Code"] = charges["Charge Code"].map(normalize_charge_code)
charges["Amount"] = pd.to_numeric(charges["Amount"], errors="coerce")
charges = charges[charges["Charge Code"].notna() & charges["Amount"].notna()].copy()

charges_wide = (
    charges.pivot_table(index="Unit", columns="Charge Code", values="Amount", aggfunc="sum", fill_value=0)
    .reset_index()
)
charges_wide.columns.name = None

out = unit_meta.merge(charges_wide, on="Unit", how="left")

known_existing = [c for c in KNOWN_CHARGE_ORDER if c in out.columns]
extra_charge_cols = sorted([c for c in charges_wide.columns if c not in {"Unit", *known_existing}])

for col in known_existing + extra_charge_cols:
    out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)

all_charge_cols = known_existing + extra_charge_cols
out["Total Charges"] = out[all_charge_cols].sum(axis=1)

output_columns = [
    "Unit",
    "Unit Type",
    "Sq Ft",
    "Resident",
    "Name",
    "Market",
    *KNOWN_CHARGE_ORDER,
    *extra_charge_cols,
    "Total Charges",
    "Resident Deposit",
    "Other Deposit",
    "Move In",
    "Lease Expiration",
    "Move Out",
    "Balance",
]

for col in output_columns:
    if col not in out.columns:
        out[col] = 0.0 if col in KNOWN_CHARGE_ORDER or col in extra_charge_cols else pd.NA

out = out[output_columns]
out = out.sort_values(by="Unit", key=lambda s: s.astype(str)).reset_index(drop=True)

out.to_excel(OUTPUT_FILE, index=False)
format_excel(OUTPUT_FILE, date_columns=["Move In", "Lease Expiration", "Move Out"])

print(f"Header row detected at Excel row {header_row + 1}")
print("Saved:", OUTPUT_FILE)
