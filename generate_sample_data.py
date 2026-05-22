"""
Generates a realistic but entirely fake rent roll Excel file for demo purposes.
Run this once to create sample_rent_roll.xlsx, then run clean_rent_roll.py.
"""

from datetime import date
import openpyxl
from openpyxl import Workbook

OUTPUT = "sample_rent_roll.xlsx"

UNITS = [
    {
        "unit": "101", "type": "1BR/1BA", "sqft": 720,
        "resident": "R1001", "name": "Alex Johnson",
        "market": 1350, "move_in": date(2023, 3, 1), "lease_exp": date(2025, 2, 28),
        "move_out": None, "balance": 0,
        "deposit": 1350, "other_deposit": 0,
        "charges": [("rrent", 1295), ("outpark", 50), ("vwaste", 18), ("pestctrl", 5)],
    },
    {
        "unit": "102", "type": "1BR/1BA", "sqft": 720,
        "resident": "R1002", "name": "Maria Garcia",
        "market": 1350, "move_in": date(2024, 1, 15), "lease_exp": date(2025, 1, 14),
        "move_out": None, "balance": 0,
        "deposit": 1350, "other_deposit": 0,
        "charges": [("rent", 1350), ("petrent", 50), ("vwaste", 18), ("pestctrl", 5)],
    },
    {
        "unit": "201", "type": "2BR/1BA", "sqft": 950,
        "resident": "R2001", "name": "David Kim",
        "market": 1650, "move_in": date(2022, 9, 1), "lease_exp": date(2024, 8, 31),
        "move_out": None, "balance": 0,
        "deposit": 1650, "other_deposit": 200,
        "charges": [("rrent", 1550), ("outpark", 50), ("outpark", 50), ("vwaste", 22), ("amenityf", 25), ("pestctrl", 5)],
    },
    {
        "unit": "202", "type": "2BR/1BA", "sqft": 950,
        "resident": "R2002", "name": "Sarah Williams",
        "market": 1650, "move_in": date(2023, 6, 1), "lease_exp": date(2025, 5, 31),
        "move_out": None, "balance": 75.50,
        "deposit": 1650, "other_deposit": 0,
        "charges": [("renl", 1625), ("petrent", 50), ("vwaste", 22), ("rcable", 55), ("pestctrl", 5)],
    },
    {
        "unit": "301", "type": "2BR/2BA", "sqft": 1100,
        "resident": "R3001", "name": "James Brown",
        "market": 1850, "move_in": date(2024, 4, 1), "lease_exp": date(2025, 3, 31),
        "move_out": None, "balance": 0,
        "deposit": 1850, "other_deposit": 0,
        "charges": [("rrent", 1800), ("outpark", 50), ("vwaste", 25), ("pestctrl", 5)],
    },
    {
        "unit": "302", "type": "2BR/2BA", "sqft": 1100,
        "resident": "R3002", "name": "Emily Chen",
        "market": 1850, "move_in": date(2023, 11, 1), "lease_exp": date(2024, 10, 31),
        "move_out": date(2024, 11, 5), "balance": -50.00,
        "deposit": 1850, "other_deposit": 0,
        "charges": [("baserent", 1850), ("vwaste", 25), ("petrent", 50), ("petrent", 50), ("pestctrl", 5)],
    },
    {
        "unit": "401", "type": "3BR/2BA", "sqft": 1350,
        "resident": "R4001", "name": "Michael Davis",
        "market": 2200, "move_in": date(2022, 7, 1), "lease_exp": date(2025, 6, 30),
        "move_out": None, "balance": 0,
        "deposit": 2200, "other_deposit": 300,
        "charges": [("rrent", 2100), ("outpark", 50), ("outpark", 50), ("vwaste", 30), ("rcable", 55), ("amenityf", 25), ("pestctrl", 5)],
    },
    {
        "unit": "402", "type": "3BR/2BA", "sqft": 1350,
        "resident": "",  "name": "",
        "market": 2200, "move_in": None, "lease_exp": None,
        "move_out": None, "balance": 0,
        "deposit": 0, "other_deposit": 0,
        "charges": [],
    },
]

wb = Workbook()
ws = wb.active
ws.title = "Rent Roll"

ws.append(["Maple Grove Apartments"])
ws.append(["Report Date:", date.today().strftime("%m/%d/%Y")])
ws.append([])

headers = [
    "Unit", "Unit Type", "Sq Ft", "Resident", "Name",
    "Market", "Charge", "Amount",
    "Move In", "Lease", "Move Out", "Balance",
    "Resident.1", "Other"
]
ws.append(headers)

for u in UNITS:
    charges = u["charges"]
    num_rows = max(len(charges), 1)

    for i in range(num_rows):
        row = [""] * len(headers)

        if i == 0:
            row[0] = u["unit"]
            row[1] = u["type"]
            row[2] = u["sqft"]
            row[3] = u["resident"]
            row[4] = u["name"]
            row[5] = u["market"]
            row[8] = u["move_in"]
            row[9] = u["lease_exp"]
            row[10] = u["move_out"]
            row[11] = u["balance"]
            row[12] = u["deposit"]
            row[13] = u["other_deposit"]

        if i < len(charges):
            row[6] = charges[i][0]
            row[7] = charges[i][1]

        ws.append(row)

    total = sum(c[1] for c in charges)
    ws.append(["", "", "", "", "", "", "Total", total])
    ws.append([])

for col_cells in ws.columns:
    max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 30)

wb.save(OUTPUT)
print(f"Sample data written to {OUTPUT}")
print(f"Units: {len([u for u in UNITS if u['charges']])}")
print("Run: python clean_rent_roll.py")
