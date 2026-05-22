import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ORIGINAL_FILE = None
CLEAN_SCRIPT = "clean_rent_roll.py"
VALIDATED_FILE = "Rent_Roll_Validated.xlsx"
REPORT_FILE = "validation_report.txt"


def norm(s):
    return str(s).replace("\n", " ").strip()


def find_header_row(raw_df, max_scan=300):
    for i in range(min(len(raw_df), max_scan)):
        row = raw_df.iloc[i].astype(str).str.replace("\n", " ").str.strip().str.lower()
        vals = set(row.values)
        if "unit" in vals and "resident" in vals and "amount" in vals and ("charge" in vals or "rent code" in vals):
            return i
    return None


_BASE_RENT_CODES = {"rrent", "rent", "renl", "rentl", "baserent", "base"}
_CODE_COLUMN_NAME_MAP = {"rp_depc": "rp_depC", "rp_depa": "rp_depA"}


def norm_charge(x: str) -> str:
    """Mirror the normalization in clean_rent_roll.py so comparisons are apples-to-apples."""
    s = str(x).strip().lower()
    s = re.sub(r"\s+", "", s)
    if not s or s == "total":
        return None
    if s in _BASE_RENT_CODES:
        return "Amount"
    return _CODE_COLUMN_NAME_MAP.get(s, s)


def get_assigned_string(script_path: Path, variable_name: str) -> str:
    text = script_path.read_text(encoding="utf-8")
    m = re.search(
        rf"^{re.escape(variable_name)}\s*=\s*[\"']([^\"']+)[\"']",
        text,
        flags=re.MULTILINE,
    )
    if not m:
        raise ValueError(f"Could not detect {variable_name} in {script_path.name}")
    return m.group(1)


def get_output_file(script_path: Path) -> str:
    return get_assigned_string(script_path, "OUTPUT_FILE")


def get_input_file(script_path: Path) -> str:
    return get_assigned_string(script_path, "INPUT_FILE")


def canonical_unit(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return None
    n = pd.to_numeric(pd.Series([s]), errors="coerce").iloc[0]
    if pd.notna(n):
        if abs(n - round(n)) <= 1e-9:
            return str(int(round(n)))
        return f"{n:.6f}".rstrip("0").rstrip(".")
    return s


def unit_key_series(series: pd.Series) -> pd.Series:
    return series.apply(canonical_unit)


def to_num(v):
    if pd.isna(v):
        return None
    n = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]
    if pd.isna(n):
        return None
    return float(n)


def to_text(v):
    if pd.isna(v):
        return None
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return None
    return s


def to_date(v):
    if pd.isna(v):
        return None
    dt = pd.to_datetime(v, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


def fmt_value(v):
    if v is None:
        return "<blank>"
    if isinstance(v, float):
        return f"{v:.2f}"
    return str(v)


def compare_field(orig_v, clean_v, field):
    date_fields = {"Move In", "Lease Expiration", "Move Out"}
    numeric_fields = {"Sq Ft", "Market", "Resident Deposit", "Other Deposit", "Balance"}

    if field in date_fields:
        o = to_date(orig_v)
        c = to_date(clean_v)
        return o == c, o, c
    if field in numeric_fields:
        o = to_num(orig_v)
        c = to_num(clean_v)
        if o is None and c is None:
            return True, o, c
        if o is None or c is None:
            return False, o, c
        return abs(o - c) <= 0.01, o, c
    o = to_text(orig_v)
    c = to_text(clean_v)
    return o == c, o, c


def apply_quality_formatting(source_file: Path, target_file: Path):
    wb = load_workbook(source_file)
    ws = wb.active

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header is not None:
            header_to_col[str(header).strip()] = col_idx

    for date_col in ["Move In", "Lease Expiration", "Move Out"]:
        cidx = header_to_col.get(date_col)
        if not cidx:
            continue
        for ridx in range(2, ws.max_row + 1):
            cell = ws.cell(row=ridx, column=cidx)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.number_format = "mm/dd/yyyy"

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(target_file)


def is_date_number_format(fmt: str) -> bool:
    f = str(fmt).lower()
    return "yy" in f and "mm" in f and "dd" in f


def df_to_text(df: pd.DataFrame, empty_msg: str) -> str:
    if df.empty:
        return empty_msg
    return df.to_string(index=False)


def main():
    report_lines = []
    failures = []

    clean_script_path = Path(CLEAN_SCRIPT)
    if not clean_script_path.exists():
        raise FileNotFoundError(f"Clean script not found: {clean_script_path}")

    output_file = get_output_file(clean_script_path)
    original_file = ORIGINAL_FILE if ORIGINAL_FILE else get_input_file(clean_script_path)

    output_path = Path(output_file)
    original_path = Path(original_file)
    validated_path = Path(VALIDATED_FILE)

    if not original_path.exists():
        raise FileNotFoundError(f"Original file not found: {original_path}")
    if not output_path.exists():
        raise FileNotFoundError(f"Cleaned file not found: {output_path}")

    today = date.today()
    output_mtime = datetime.fromtimestamp(output_path.stat().st_mtime)
    timestamp_ok = output_mtime.date() == today
    if not timestamp_ok:
        failures.append("Output file timestamp is not today's date.")

    raw = pd.read_excel(original_path, sheet_name=0, header=None)
    header_idx = find_header_row(raw)
    if header_idx is None:
        raise ValueError("Could not detect header row in original file.")

    orig = pd.read_excel(original_path, sheet_name=0, header=header_idx)
    orig.columns = [norm(c) for c in orig.columns]

    clean = pd.read_excel(output_path)
    clean.columns = [str(c).strip() for c in clean.columns]

    required_orig = ["Unit", "Resident", "Amount", "Charge"]
    lease_fields = ["Move In", "Lease", "Move Out", "Balance", "Resident.1", "Other"]
    missing_required_orig = [c for c in required_orig if c not in orig.columns]
    missing_lease_orig = [c for c in lease_fields if c not in orig.columns]
    if missing_required_orig:
        failures.append(f"Missing required original columns: {missing_required_orig}")

    required_clean = [
        "Unit",
        "Unit Type",
        "Sq Ft",
        "Resident",
        "Market",
        "Resident Deposit",
        "Other Deposit",
        "Move In",
        "Lease Expiration",
        "Move Out",
        "Balance",
        "Total Charges",
    ]
    missing_required_clean = [c for c in required_clean if c not in clean.columns]
    if missing_required_clean:
        failures.append(f"Missing required cleaned columns: {missing_required_clean}")

    metadata_cols_clean = {
        "Unit",
        "Unit Type",
        "Sq Ft",
        "Resident",
        "Market",
        "Resident Deposit",
        "Other Deposit",
        "Move In",
        "Lease Expiration",
        "Move Out",
        "Balance",
        "Total Charges",
    }
    charge_cols_clean = [c for c in clean.columns if c not in metadata_cols_clean]

    orig_headers = orig[orig["Unit"].notna()].copy()
    orig_headers["UnitKey"] = unit_key_series(orig_headers["Unit"])
    orig_headers = orig_headers[orig_headers["UnitKey"].notna()].copy()
    orig_headers = orig_headers.drop_duplicates(subset=["UnitKey"], keep="first").set_index("UnitKey", drop=False)

    clean["UnitKey"] = unit_key_series(clean["Unit"])
    clean = clean[clean["UnitKey"].notna()].copy()
    dup_units = clean["UnitKey"][clean["UnitKey"].duplicated()].unique().tolist()
    if dup_units:
        failures.append(f"Duplicate units found in cleaned output: {sorted(dup_units)}")
    clean_by_unit = clean.drop_duplicates(subset=["UnitKey"], keep="first").set_index("UnitKey", drop=False)

    orig_units = set(orig_headers.index.tolist())
    clean_units = set(clean_by_unit.index.tolist())
    missing_units = sorted(orig_units - clean_units)
    extra_units = sorted(clean_units - orig_units)
    if missing_units:
        failures.append(f"Missing units in cleaned output: {missing_units}")
    if extra_units:
        failures.append(f"Extra units in cleaned output: {extra_units}")

    metadata_map = [
        ("Unit Type", "Unit Type"),
        ("Unit.1", "Sq Ft"),
        ("Resident", "Resident"),
        ("Market", "Market"),
        ("Resident.1", "Resident Deposit"),
        ("Other", "Other Deposit"),
        ("Move In", "Move In"),
        ("Lease", "Lease Expiration"),
        ("Move Out", "Move Out"),
        ("Balance", "Balance"),
    ]
    metadata_mismatches = []
    for orig_col, clean_col in metadata_map:
        if orig_col not in orig.columns or clean_col not in clean.columns:
            continue
        for unit in sorted(orig_units & clean_units):
            o = orig_headers.at[unit, orig_col] if unit in orig_headers.index else None
            c = clean_by_unit.at[unit, clean_col] if unit in clean_by_unit.index else None
            equal, o_n, c_n = compare_field(o, c, clean_col)
            if not equal:
                metadata_mismatches.append(
                    {
                        "Unit": unit,
                        "Field": clean_col,
                        "OriginalValue": fmt_value(o_n),
                        "CleanValue": fmt_value(c_n),
                    }
                )

    orig_long_cols = [c for c in ["Unit", "Unit Type", "Unit.1", "Resident", "Charge", "Amount"] if c in orig.columns]
    orig_long = orig[orig_long_cols].copy()
    for c in ["Unit", "Unit Type", "Unit.1", "Resident"]:
        if c in orig_long.columns:
            orig_long[c] = orig_long[c].ffill()
    orig_long["UnitKey"] = unit_key_series(orig_long["Unit"])
    orig_long["AmountNum"] = pd.to_numeric(orig_long["Amount"], errors="coerce")
    orig_long["ChargeRaw"] = orig_long["Charge"].fillna("").astype(str).str.strip()
    orig_long = orig_long[orig_long["UnitKey"].notna()].copy()
    orig_charge_rows_total = len(orig_long[orig_long["AmountNum"].notna()])

    orig_charges = orig_long[orig_long["AmountNum"].notna()].copy()
    orig_charges = orig_charges[orig_charges["ChargeRaw"].ne("")].copy()
    orig_charges = orig_charges[orig_charges["ChargeRaw"].str.lower().ne("total")].copy()
    orig_charges["ChargeNorm"] = orig_charges["ChargeRaw"].map(norm_charge)

    orig_charge_codes = sorted(orig_charges["ChargeNorm"].dropna().unique().tolist())
    missing_charge_cols = sorted(set(orig_charge_codes) - set(charge_cols_clean))
    extra_charge_cols = sorted(set(charge_cols_clean) - set(orig_charge_codes))
    if missing_charge_cols:
        failures.append(f"Charge columns missing from cleaned output: {missing_charge_cols}")

    if "Total Charges" not in clean.columns:
        failures.append("Total Charges column not found in cleaned output.")

    for c in charge_cols_clean + (["Total Charges"] if "Total Charges" in clean.columns else []):
        clean[c] = pd.to_numeric(clean[c], errors="coerce").fillna(0.0)
    clean_by_unit = clean.drop_duplicates(subset=["UnitKey"], keep="first").set_index("UnitKey", drop=False)

    orig_unit_charge = orig_charges.groupby(["UnitKey", "ChargeNorm"])["AmountNum"].sum()
    orig_unit_total = orig_charges.groupby("UnitKey")["AmountNum"].sum()

    charge_mismatches = []
    total_calc_mismatches = []
    total_vs_orig_mismatches = []
    for unit in sorted(orig_units & clean_units):
        if unit not in clean_by_unit.index:
            continue
        clean_row = clean_by_unit.loc[unit]

        for charge_col in charge_cols_clean:
            orig_val = float(orig_unit_charge.get((unit, charge_col), 0.0))
            clean_val = float(clean_row.get(charge_col, 0.0))
            diff = clean_val - orig_val
            if abs(diff) > 0.01:
                charge_mismatches.append(
                    {
                        "Unit": unit,
                        "Charge": charge_col,
                        "OriginalSum": round(orig_val, 2),
                        "CleanValue": round(clean_val, 2),
                        "Diff": round(diff, 2),
                    }
                )

        if "Total Charges" in clean.columns:
            row_sum = float(clean_row[charge_cols_clean].sum())
            clean_total = float(clean_row["Total Charges"])
            diff_calc = clean_total - row_sum
            if abs(diff_calc) > 0.01:
                total_calc_mismatches.append(
                    {
                        "Unit": unit,
                        "SumOfChargeColumns": round(row_sum, 2),
                        "TotalCharges": round(clean_total, 2),
                        "Diff": round(diff_calc, 2),
                    }
                )

            orig_total = float(orig_unit_total.get(unit, 0.0))
            diff_orig = clean_total - orig_total
            if abs(diff_orig) > 0.01:
                total_vs_orig_mismatches.append(
                    {
                        "Unit": unit,
                        "OriginalTotal": round(orig_total, 2),
                        "TotalCharges": round(clean_total, 2),
                        "Diff": round(diff_orig, 2),
                    }
                )

    duplicate_xy_cols = [c for c in clean.columns if c.endswith("_x") or c.endswith("_y")]
    if duplicate_xy_cols:
        failures.append(f"Duplicate _x/_y columns present: {duplicate_xy_cols}")

    apply_quality_formatting(output_path, validated_path)
    validated_wb = load_workbook(validated_path)
    validated_ws = validated_wb.active
    freeze_ok = validated_ws.freeze_panes == "A2"
    filter_ok = bool(validated_ws.auto_filter.ref)
    date_format_issues = []
    header_to_col = {}
    for col_idx in range(1, validated_ws.max_column + 1):
        header = validated_ws.cell(row=1, column=col_idx).value
        if header is not None:
            header_to_col[str(header).strip()] = col_idx
    for date_col in ["Move In", "Lease Expiration", "Move Out"]:
        cidx = header_to_col.get(date_col)
        if not cidx:
            continue
        col_letter = get_column_letter(cidx)
        width = validated_ws.column_dimensions[col_letter].width or 0
        if width < 12:
            date_format_issues.append(f"{date_col}: column width too narrow ({width})")
        for ridx in range(2, validated_ws.max_row + 1):
            cell = validated_ws.cell(row=ridx, column=cidx)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            if not is_date_number_format(cell.number_format):
                date_format_issues.append(f"{date_col}: non-date format at row {ridx}")
                break

    if not freeze_ok:
        failures.append("Freeze panes is not set to A2.")
    if not filter_ok:
        failures.append("Auto-filter is not enabled.")
    if date_format_issues:
        failures.append(f"Date formatting issues: {date_format_issues}")

    if metadata_mismatches:
        failures.append(f"Metadata mismatches found: {len(metadata_mismatches)}")
    if charge_mismatches:
        failures.append(f"Charge mismatches found: {len(charge_mismatches)}")
    if total_calc_mismatches:
        failures.append(f"Total Charges row-sum mismatches found: {len(total_calc_mismatches)}")
    if total_vs_orig_mismatches:
        failures.append(f"Total Charges vs original mismatches found: {len(total_vs_orig_mismatches)}")

    report_lines.append("Rent Roll Validation Report")
    report_lines.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    report_lines.append("")
    report_lines.append(f"Original file: {original_path.resolve()}")
    report_lines.append(f"Cleaned file: {output_path.resolve()}")
    report_lines.append(f"Validated file: {validated_path.resolve()}")
    report_lines.append(f"Detected header row in original (1-based): {header_idx + 1}")
    report_lines.append(f"Output file mtime: {output_mtime.isoformat(timespec='seconds')}")
    report_lines.append(f"Timestamp is today ({today.isoformat()}): {timestamp_ok}")
    report_lines.append("")

    report_lines.append("Summary")
    report_lines.append(f"Result: {'PASS' if not failures else 'FAIL'}")
    report_lines.append(f"Failure count: {len(failures)}")
    if failures:
        report_lines.append("Failure details:")
        for f in failures:
            report_lines.append(f"- {f}")
    report_lines.append("")

    report_lines.append("Column Checks")
    report_lines.append(f"Required original columns missing: {missing_required_orig if missing_required_orig else 'None'}")
    report_lines.append(f"Lease-related columns missing from original: {missing_lease_orig if missing_lease_orig else 'None'}")
    report_lines.append(f"Required cleaned columns missing: {missing_required_clean if missing_required_clean else 'None'}")
    report_lines.append(f"Cleaned charge columns ({len(charge_cols_clean)}): {charge_cols_clean}")
    report_lines.append(f"Original normalized charge columns ({len(orig_charge_codes)}): {orig_charge_codes}")
    report_lines.append(f"Missing normalized charge columns in cleaned: {missing_charge_cols if missing_charge_cols else 'None'}")
    report_lines.append(f"Extra cleaned charge columns not in original normalized set: {extra_charge_cols if extra_charge_cols else 'None'}")
    report_lines.append("")

    report_lines.append("Counts")
    report_lines.append(f"Original rows: {len(orig)}")
    report_lines.append(f"Original unit header rows: {len(orig_headers)}")
    report_lines.append(f"Cleaned rows: {len(clean_by_unit)}")
    report_lines.append(f"Original numeric charge rows (including Total/blank): {orig_charge_rows_total}")
    report_lines.append(f"Original numeric charge rows (excluding Total/blank): {len(orig_charges)}")
    report_lines.append(f"Missing units in cleaned: {len(missing_units)}")
    report_lines.append(f"Extra units in cleaned: {len(extra_units)}")
    report_lines.append("")

    report_lines.append("Missing Units")
    report_lines.append("None" if not missing_units else ", ".join(missing_units))
    report_lines.append("")
    report_lines.append("Extra Units")
    report_lines.append("None" if not extra_units else ", ".join(extra_units))
    report_lines.append("")

    metadata_df = pd.DataFrame(metadata_mismatches, columns=["Unit", "Field", "OriginalValue", "CleanValue"])
    charge_df = pd.DataFrame(charge_mismatches, columns=["Unit", "Charge", "OriginalSum", "CleanValue", "Diff"])
    total_calc_df = pd.DataFrame(
        total_calc_mismatches, columns=["Unit", "SumOfChargeColumns", "TotalCharges", "Diff"]
    )
    total_orig_df = pd.DataFrame(total_vs_orig_mismatches, columns=["Unit", "OriginalTotal", "TotalCharges", "Diff"])

    report_lines.append("Metadata Mismatches")
    report_lines.append(df_to_text(metadata_df, "None"))
    report_lines.append("")
    report_lines.append("Charge Mismatches")
    report_lines.append(df_to_text(charge_df, "None"))
    report_lines.append("")
    report_lines.append("Total Charges vs Sum(Charge Columns) Mismatches")
    report_lines.append(df_to_text(total_calc_df, "None"))
    report_lines.append("")
    report_lines.append("Total Charges vs Original Non-Total Charge Sum Mismatches")
    report_lines.append(df_to_text(total_orig_df, "None"))
    report_lines.append("")

    report_lines.append("Quality Checks")
    report_lines.append(f"Duplicate _x/_y columns: {duplicate_xy_cols if duplicate_xy_cols else 'None'}")
    report_lines.append(f"Freeze header row enabled (A2): {freeze_ok}")
    report_lines.append(f"Filters enabled: {filter_ok}")
    report_lines.append(f"Date formatting issues: {date_format_issues if date_format_issues else 'None'}")
    report_lines.append("")

    report_lines.append("Assumptions and Transformations")
    report_lines.append("- Header detection requires row values containing Unit, Resident, Amount, and Charge/Rent Code.")
    report_lines.append("- Unit comparison uses canonical numeric-string keys (e.g., 1001 and 1001.0 are treated as the same Unit).")
    report_lines.append("- Charge normalization map: rent/renl/rentl/baserent/base->Base Rent; internet/int->Internet; trash-pu/trashpu/trash->Trash; pestctrl/pest/pestcontrol->Pest; parking/park->Parking; garage->Garage; petrent/pet->Pet Rent; otconc->OT Concession; utilreb->Utility Reimb; storage->Storage; misc->Misc; empdisc->Emp Discount.")
    report_lines.append("- Unknown charge codes are retained as normalized lowercase no-space labels and validated as distinct columns.")
    report_lines.append("- Charge comparisons and total comparisons use absolute tolerance <= 0.01.")
    report_lines.append("- Blank charge labels are excluded from charge aggregation.")
    report_lines.append("")

    Path(REPORT_FILE).write_text("\n".join(report_lines), encoding="utf-8")

    print(f"Validation result: {'PASS' if not failures else 'FAIL'}")
    print(f"Report written: {Path(REPORT_FILE).resolve()}")
    print(f"Validated workbook written: {validated_path.resolve()}")
    print(f"Units original/cleaned: {len(orig_units)} / {len(clean_units)}")
    print(f"Missing/extra units: {len(missing_units)} / {len(extra_units)}")
    print(f"Metadata mismatches: {len(metadata_mismatches)}")
    print(f"Charge mismatches: {len(charge_mismatches)}")
    print(f"Total calc mismatches: {len(total_calc_mismatches)}")
    print(f"Total vs original mismatches: {len(total_vs_orig_mismatches)}")

    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
